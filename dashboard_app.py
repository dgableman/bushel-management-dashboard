"""
Standalone Bushel Management Dashboard Application
Run this script to launch the interactive dashboard in a web browser.
"""

import sys
import os
from pathlib import Path
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from datetime import datetime, date
import streamlit as st
import yfinance as yf

# Detect environment (Colab, Streamlit Cloud, or local)
def is_colab():
    """Check if running in Google Colab."""
    try:
        import google.colab
        return True
    except ImportError:
        return False

def is_streamlit_cloud():
    """Check if running on Streamlit Cloud."""
    # Streamlit Cloud sets these environment variables or paths
    script_path = str(Path(__file__).absolute())
    return (os.getenv('STREAMLIT_SERVER_PORT') is not None or 
            os.getenv('STREAMLIT_SHARING_MODE') is not None or
            '/mount/src' in script_path or
            'streamlit.app' in os.getenv('STREAMLIT_SERVER_ADDRESS', ''))

# Set up paths based on environment
if is_colab():
    # Colab paths
    PROJECT_PATH = os.getenv('PROJECT_PATH', '/content/drive/MyDrive/Colab_Notebooks/Grain_Manager')
    DB_PATH = os.getenv('DB_PATH', '/content/drive/MyDrive/Colab_Notebooks/Grain_Manager/database/bushel_management.db')
elif is_streamlit_cloud():
    # Streamlit Cloud paths
    # Streamlit Cloud mounts the repo at /mount/src/REPO_NAME
    SCRIPT_DIR = Path(__file__).parent.absolute()
    PROJECT_PATH = str(SCRIPT_DIR)
    
    # Try Streamlit secrets first (recommended for production)
    # Note: st.secrets may not be available at import time, so we'll check it in the main function
    DB_PATH = os.getenv('DB_PATH')  # Check environment variable first
    
    # If not set, try to find the database file
    if not DB_PATH:
        # Try multiple possible paths for the database
        possible_paths = [
            str(SCRIPT_DIR / 'data' / 'bushel_management.db'),  # Standard path
            str(SCRIPT_DIR / 'data' / 'bushel-management.db'),   # Hyphen version
            '/mount/src/bushel-management-dashboard/data/bushel_management.db',  # Absolute Streamlit Cloud path
            '/mount/src/bushel-management-dashboard/data/bushel-management.db',  # Absolute with hyphen
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                DB_PATH = path
                break
        else:
            # Default to the most likely path
            DB_PATH = possible_paths[0]
else:
    # Local paths - adjust these to match your local setup
    # Get the directory where this script is located
    SCRIPT_DIR = Path(__file__).parent.absolute()
    PROJECT_PATH = str(SCRIPT_DIR)
    # Default local database path - use the main Bushel_Management project database
    # This allows the reporting project to always use the most up-to-date database
    default_local_db = Path.home() / 'PycharmProjects' / 'Bushel_Management' / 'data' / 'bushel_management.db'
    # Fallback to local data folder if main project database doesn't exist
    if not default_local_db.exists():
        default_local_db = SCRIPT_DIR / 'data' / 'bushel_management.db'
    DB_PATH = os.getenv('DB_PATH', str(default_local_db))

# Add project to path
if PROJECT_PATH not in sys.path:
    sys.path.insert(0, PROJECT_PATH)

from database.db_connection import create_db_session
from reports.contract_queries import get_all_contracts, get_active_contracts
from reports.settlement_queries import get_all_settlements
from reports.bin_queries import (
    get_bins_with_storage_by_crop,
    get_bin_storage_metrics,
    build_open_contract_allocation_by_bin,
)
from reports.commodity_utils import (
    normalize_commodity_name,
    get_commodities_for_normalized_name,
    get_all_normalized_commodities
)
from reports.vendor_utils import normalize_vendor_name, get_all_normalized_vendors
from reports.crop_year_utils import (
    get_current_crop_year,
    get_crop_year_date_range,
    get_crop_year_from_date,
    get_display_year_options,
    get_display_calendar_year_options,
    discover_contract_crop_years,
    discover_contract_calendar_years,
    delivery_months_for_year_selection,
    contract_matches_year_basis,
    format_delivery_month_key,
)
from reports.crop_year_sales import calculate_crop_year_sales
from reports.monthly_deliveries import (
    calculate_monthly_deliveries,
    get_month_name_for_crop_year
)
from reports.contract_pdf_storage import (
    fetch_pdf_bytes,
    storage_available,
    list_available_contract_numbers,
    signed_url_for_contract,
)

# Set page config for full-width layout
st.set_page_config(
    page_title="Bushel Management Dashboard",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded"
)

def get_drilldown_details(
    db,
    crop: str,
    status: str,
    crop_year: int,
    all_contracts,
    all_settlements
):
    """
    Get detailed contract and settlement data for drill-down view.
    
    Args:
        db: Database session
        crop: Normalized crop name
        status: Status to filter ('Sold', 'Contracted', 'Open')
        crop_year: Crop year
        all_contracts: List of all contracts
        all_settlements: List of all settlements
        
    Returns:
        Dictionary with 'contracts' and 'settlements' lists
    """
    from reports.crop_year_utils import (
        is_date_in_crop_year, 
        calculate_settlement_revenue, 
        calculate_partial_contract_remaining,
        get_starting_bushels,
        get_crop_year_date_range
    )
    from reports.commodity_utils import normalize_commodity_name, get_commodities_for_normalized_name
    
    start_date, end_date = get_crop_year_date_range(crop_year)
    
    # Get commodities that map to this normalized crop
    crop_aliases = get_commodities_for_normalized_name(db, crop)
    
    details = {
        'contracts': [],
        'settlements': []
    }
    
    if status == 'Sold':
        # Show all settlement lines (not just headers) for this crop and crop year
        # Group by settlement ID and contract ID, sum bushels and amounts
        settlement_groups = {}  # Key: (settlement_ID, contract_id), Value: aggregated data
        
        for settlement in all_settlements:
            # Skip header rows - we want the actual line items
            if settlement.status == 'Header':
                continue
            if not settlement.date_delivered or not is_date_in_crop_year(settlement.date_delivered, crop_year):
                continue
            if not settlement.commodity:
                continue
            normalized = normalize_commodity_name(db, settlement.commodity)
            if normalized != crop:
                continue
            
            # Get contract ID - if None, empty, or "none", it's an Open Sale
            contract_id = settlement.contract_id
            if contract_id is None:
                contract_id = 'none'
            elif isinstance(contract_id, str) and contract_id.strip().lower() in ['none', '', 'null']:
                contract_id = 'none'
            else:
                contract_id = str(contract_id).strip()
            
            settlement_id = settlement.settlement_ID
            
            # Create key for grouping (settlement_ID, contract_id)
            group_key = (settlement_id, contract_id)
            
            if group_key not in settlement_groups:
                settlement_groups[group_key] = {
                    'settlement_id': settlement_id,
                    'contract_id': 'Open Sale' if contract_id == 'none' else contract_id,
                    'date_delivered': settlement.date_delivered,
                    'bushels': 0,
                    'price': settlement.price,
                    'gross_amount': 0.0,
                    'net_amount': 0.0,
                    'buyer': settlement.buyer,
                    'commodity': settlement.commodity
                }
            
            # Sum bushels and amounts for this group
            settlement_groups[group_key]['bushels'] += (settlement.bushels or 0)
            settlement_groups[group_key]['gross_amount'] += (settlement.gross_amount or 0.0)
            # For net_amount, prefer net_amount, fall back to gross_amount
            if settlement.net_amount:
                settlement_groups[group_key]['net_amount'] += settlement.net_amount
            elif settlement.gross_amount:
                settlement_groups[group_key]['net_amount'] += settlement.gross_amount
        
        # Convert grouped data to list
        for group_data in settlement_groups.values():
            details['settlements'].append(group_data)
    
    elif status == 'Contracted':
        # Show active open or partial contracts for this crop and crop year
        from reports.crop_year_utils import calculate_partial_contract_remaining
        
        for contract in all_contracts:
            # Filter by status='Active'
            contract_status = (contract.status or '').strip()
            if contract_status.lower() != 'active':
                continue
            
            if not contract.delivery_start or not is_date_in_crop_year(contract.delivery_start, crop_year):
                continue
            if not contract.commodity:
                continue
            normalized = normalize_commodity_name(db, contract.commodity)
            if normalized != crop:
                continue
            
            fill_status = contract.fill_status or 'None'
            if fill_status not in ['None', 'Partial']:
                continue
            
            # Calculate remaining if partial
            if fill_status == 'None':
                remaining_bushels = contract.bushels or 0
                remaining_revenue = remaining_bushels * (contract.price or 0.0)
            else:
                remaining_revenue, remaining_bushels = calculate_partial_contract_remaining(
                    contract, all_settlements
                )
            
            if remaining_bushels > 0:
                details['contracts'].append({
                    'contract_number': contract.contract_number,
                    'commodity': contract.commodity,
                    'bushels': contract.bushels or 0,
                    'remaining_bushels': remaining_bushels,
                    'price': contract.price,
                    'remaining_revenue': remaining_revenue,
                    'fill_status': fill_status,
                    'delivery_start': contract.delivery_start,
                    'buyer': contract.buyer_name,
                    'date_sold': contract.date_sold
                })
    
    elif status == 'Open':
        # Calculate Open bushels: Starting - Sold - Contracted
        from reports.crop_year_utils import get_starting_bushels, is_date_in_crop_year, calculate_partial_contract_remaining
        
        starting_bushels = get_starting_bushels(db, crop_year, crop)
        
        # Calculate sold bushels (all settlement header rows for this crop/year)
        # Use same logic as calculate_crop_year_sales
        sold_bushels = 0
        for settlement in all_settlements:
            # Check status - handle case-insensitive and various formats
            status = (settlement.status or '').strip()
            if status.lower() != 'header':
                continue
            if not settlement.date_delivered or not is_date_in_crop_year(settlement.date_delivered, crop_year):
                continue
            if not settlement.commodity:
                continue
            normalized = normalize_commodity_name(db, settlement.commodity)
            if normalized == crop:
                if settlement.bushels:
                    sold_bushels += settlement.bushels
        
        # Calculate contracted bushels (remaining bushels on active open/partial contracts for this crop/year)
        contracted_bushels = 0
        for contract in all_contracts:
            # Filter by status='Active'
            contract_status = (contract.status or '').strip()
            if contract_status.lower() != 'active':
                continue
            
            if not contract.delivery_start or not is_date_in_crop_year(contract.delivery_start, crop_year):
                continue
            if not contract.commodity:
                continue
            normalized = normalize_commodity_name(db, contract.commodity)
            if normalized != crop:
                continue
            
            fill_status = contract.fill_status or 'None'
            if fill_status not in ['None', 'Partial']:
                continue
            
            if fill_status == 'None':
                remaining_bushels = contract.bushels or 0
            else:
                _, remaining_bushels = calculate_partial_contract_remaining(contract, all_settlements)
            
            if remaining_bushels > 0:
                contracted_bushels += remaining_bushels
        
        open_bushels = max(0, starting_bushels - sold_bushels - contracted_bushels)
        
        details['summary'] = {
            'starting_bushels': starting_bushels,
            'sold_bushels': sold_bushels,
            'contracted_bushels': contracted_bushels,
            'open_bushels': open_bushels
        }
    
    return details


BIN_SETTLED_COLOR = '#2ecc71'
BIN_SETTLED_LINE_COLOR = '#1e8449'
BIN_CONTRACTED_COLOR = '#9b59b6'
BIN_CAPTION_STYLE = (
    "text-align:center;font-size:1.35rem;font-weight:700;"
    "font-family:Arial Black,sans-serif;margin:0;padding:0;line-height:1.2;"
)
BIN_CAPTION_WRAP = "margin-top:-1.25rem;margin-bottom:0.25rem;"


def _short_bin_title(label: str) -> str:
    title = label.split(' - ', 1)[-1] if ' - ' in label else label
    return title[:28] + ('…' if len(title) > 28 else '')


def _bin_metrics_label_text(metrics: dict) -> str:
    """Top-of-bar metrics (availability shown separately below the chart)."""
    if metrics['is_unlimited']:
        cap_line = f"Current: {metrics['current']:,.0f} / {metrics['reference']:,.0f} bu (unlimited)"
    elif metrics['capacity'] > 0:
        cap_line = f"Current: {metrics['current']:,.0f} / {metrics['capacity']:,.0f} bu"
    else:
        cap_line = f"Current: {metrics['current']:,.0f} bu"
    contracted_raw = metrics.get('open_contracts', metrics.get('contracted_raw', metrics['contracted']))
    not_sold = metrics.get('not_sold', metrics.get('available_to_market', 0))
    return (
        f"Initial: {metrics['initial']:,.0f} bu<br>"
        f"Open Contracts: {contracted_raw:,.0f} bu<br>"
        f"Settled: {metrics['settled']:,.0f} bu<br>"
        f"Not Sold: {not_sold:,.0f} bu<br>"
        f"{cap_line}"
    )


def _bin_bar_label_yshifts(metrics_html: str) -> tuple:
    """Return (bin_name_yshift, metrics_yshift) so the title sits above the metrics block."""
    metrics_lines = metrics_html.count('<br>') + 1
    line_height = 17
    metrics_yshift = 12
    name_yshift = metrics_yshift + (metrics_lines * line_height) + 14
    return name_yshift, metrics_yshift


def _availability_caption_text(label: str) -> str:
    return label.replace('\n', ' — ')


def _render_bin_availability_captions(availability_labels: list):
    if not availability_labels:
        return
    if len(availability_labels) == 1:
        st.markdown(
            f"<div style='{BIN_CAPTION_WRAP}'><p style='{BIN_CAPTION_STYLE}'>"
            f"{_availability_caption_text(availability_labels[0])}</p></div>",
            unsafe_allow_html=True,
        )
        return
    cap_cols = st.columns(len(availability_labels))
    for cap_col, avail_text in zip(cap_cols, availability_labels):
        cap_col.markdown(
            f"<div style='{BIN_CAPTION_WRAP}'><p style='{BIN_CAPTION_STYLE}'>"
            f"{_availability_caption_text(avail_text)}</p></div>",
            unsafe_allow_html=True,
        )


def add_bins_stacked_bar_traces(
    fig,
    x_labels,
    settled,
    contracted,
    uncontracted,
    empty_space,
    crop_colors,
    crop_line_colors,
    empty_colors,
    bar_width,
    show_legend=True,
    legend_suffix='',
    subplot_row=None,
    subplot_col=None,
):
    """Stacked bar segments: Settled, Open Contracts, Not sold in-bin, then Empty."""
    width = [bar_width] * len(x_labels)
    traces = [
        ('Settled', settled, BIN_SETTLED_COLOR, BIN_SETTLED_LINE_COLOR, 'Settled: %{y:,.0f} bu'),
        ('Open Contracts', contracted, BIN_CONTRACTED_COLOR, '#7d3c98', 'Open contracts: %{y:,.0f} bu'),
        ('Not sold', uncontracted, crop_colors, crop_line_colors, 'Not sold: %{y:,.0f} bu'),
        ('Empty', empty_space, empty_colors, empty_colors, 'Empty: %{y:,.0f} bu'),
    ]
    for idx, (name, y, color, line_color, hover_key) in enumerate(traces):
        legend_name = f'{legend_suffix}{name}'.strip() if legend_suffix else name
        marker = dict(
            color=color,
            line=dict(color=line_color, width=1),
            cornerradius=0.2,
        )
        trace = go.Bar(
            x=x_labels,
            y=y,
            name=legend_name,
            orientation='v',
            marker=marker,
            hovertemplate=f'<b>%{{x}}</b><br>{hover_key}<extra></extra>',
            showlegend=show_legend,
            width=width,
            text=[''] * len(x_labels),
        )
        if subplot_row is not None:
            fig.add_trace(trace, row=subplot_row, col=subplot_col)
        else:
            fig.add_trace(trace)


@st.cache_resource
def get_database_session():
    """Create and cache a database session using the configured DB_PATH."""
    try:
        db_path = DB_PATH

        # Allow a Streamlit secret or env override (used for cloud deployments)
        try:
            if hasattr(st, 'secrets') and st.secrets.get("DB_PATH"):
                db_path = st.secrets.get("DB_PATH")
        except (AttributeError, FileNotFoundError, KeyError):
            pass

        if not Path(db_path).exists():
            raise FileNotFoundError(f"Database file not found: {db_path}")

        return create_db_session(db_path)
    except FileNotFoundError as e:
        st.error(f"❌ {e}")
        return None
    except Exception as e:
        st.error(f"❌ Database connection error: {type(e).__name__}: {e}")
        import traceback
        with st.expander("🔍 Full error details"):
            st.code(traceback.format_exc())
        return None

@st.cache_data(ttl=3600)  # Cache for 1 hour
def get_market_prices():
    """
    Fetch current market prices from Yahoo Finance.
    Returns default prices with fallback if API fails.
    Futures prices are in cents, so divide by 100 to get dollars per bushel.
    """
    default_prices = {
        'Corn': 4.0,
        'Soybeans': 10.0
    }
    
    try:
        # Fetch corn futures price (ZC=F)
        corn_ticker = yf.Ticker("ZC=F")
        corn_data = corn_ticker.history(period="1d")
        if not corn_data.empty:
            corn_price_cents = float(corn_data['Close'].iloc[-1])
            corn_price_dollars = corn_price_cents / 100.0
            default_prices['Corn'] = max(0.0, corn_price_dollars - 0.30)  # Divide by 100, then subtract 30 cents
        
        # Fetch soybean futures price (ZS=F)
        bean_ticker = yf.Ticker("ZS=F")
        bean_data = bean_ticker.history(period="1d")
        if not bean_data.empty:
            bean_price_cents = float(bean_data['Close'].iloc[-1])
            bean_price_dollars = bean_price_cents / 100.0
            default_prices['Soybeans'] = max(0.0, bean_price_dollars - 0.45)  # Divide by 100, then subtract 45 cents
    except Exception as e:
        # If yfinance fails, use default values
        # Silently fail and use hardcoded defaults
        pass
    
    return default_prices


def render_contract_pdf_picker(contract_numbers, key_prefix, label="\U0001F4C4 View contract PDF"):
    """Render a contract picker + Download PDF button.

    PDFs live in a private Google Cloud Storage bucket and are fetched lazily
    (and cached) only for the contract the user selects.
    """
    numbers = sorted({(n or "").strip() for n in contract_numbers if (n or "").strip()})
    if not numbers:
        return

    st.markdown(f"##### {label}")

    if not storage_available():
        st.info(
            "Contract PDFs aren't configured yet. Add the `gcp_service_account` "
            "secret in the app settings to enable PDF downloads."
        )
        return

    selected = st.selectbox(
        "Contract number",
        options=["\u2014 Select \u2014"] + numbers,
        key=f"{key_prefix}_pdf_select",
    )
    if selected and selected != "\u2014 Select \u2014":
        pdf_bytes = fetch_pdf_bytes(selected)
        if pdf_bytes:
            st.download_button(
                label=f"\u2B07\uFE0F Download {selected}.pdf",
                data=pdf_bytes,
                file_name=f"{selected}.pdf",
                mime="application/pdf",
                key=f"{key_prefix}_pdf_download",
            )
        else:
            st.warning(f"No PDF found in storage for contract {selected}.")


_DELIVERIES_FILL_STATUSES = ("None", "Partial", "Filled", "Over")
_DELIVERIES_STATUS_SORT = {"Partial": 0, "Open": 1, "Filled": 2, "Over": 3}


def _deliveries_fill_status_label(fill_status: str) -> str:
    return "Open" if fill_status == "None" else fill_status


def _contract_delivered_bushels(contract, settlements) -> int:
    contract_number = contract.contract_number
    return sum(
        s.bushels or 0
        for s in settlements
        if s.contract_id == contract_number
        and (s.status or "").strip().lower() != "header"
    )


def render_deliveries_tab(db, contracts, settlements):
    """Render the Deliveries tab.

    Lets the user pick a calendar Year and Month (NOT crop year), then shows
    contracts whose delivery window overlaps that month and whose fill status is
    Open, Partial, Filled, or Over, grouped by commodity (normalized) and
    location (buyer name):
      - a Summary table: commodity, location, total/delivered/remaining bushels
      - per commodity+location: each contract's number, status, price, and bushels
    """
    import calendar

    st.subheader("🚚 Deliveries")

    today = datetime.now()
    col_year, col_month, _spacer = st.columns([1, 1, 2])
    with col_year:
        year_options = get_display_calendar_year_options(today.year)
        selected_year = st.selectbox(
            "Year",
            options=year_options,
            index=(
                year_options.index(today.year)
                if today.year in year_options
                else len(year_options) - 1
            ),
            key="deliveries_tab_year",
        )
    with col_month:
        selected_month = st.selectbox(
            "Month",
            options=list(range(1, 13)),
            index=today.month - 1,
            format_func=lambda m: calendar.month_name[m],
            key="deliveries_tab_month",
        )

    # Calendar-month window (inclusive of both edges)
    start_of_month = date(selected_year, selected_month, 1)
    last_day = calendar.monthrange(selected_year, selected_month)[1]
    end_of_month = date(selected_year, selected_month, last_day)

    st.caption(
        f"Contracts delivering in {calendar.month_name[selected_month]} {selected_year} "
        f"({start_of_month:%m/%d/%Y} \u2013 {end_of_month:%m/%d/%Y})"
    )

    # Filter to contracts whose delivery window overlaps the selected month and
    # whose fill status is Open/Partial/Filled/Over, then group by (commodity, location).
    groups = {}  # (commodity, location) -> totals + detail rows
    for c in contracts:
        fill_status = c.fill_status or "None"
        if fill_status not in _DELIVERIES_FILL_STATUSES:
            continue

        eff_start = c.delivery_start or c.delivery_end
        eff_end = c.delivery_end or c.delivery_start
        if not eff_start or not eff_end:
            continue
        if eff_start > eff_end:  # guard against reversed dates
            eff_start, eff_end = eff_end, eff_start

        # Overlap test: window touches the month if it starts on/before month end
        # and ends on/after month start.
        if not (eff_start <= end_of_month and eff_end >= start_of_month):
            continue

        commodity = normalize_commodity_name(db, c.commodity)
        location = normalize_vendor_name(db, c.buyer_name)  # roll up vendor spellings
        bushels = c.bushels or 0
        delivered = _contract_delivered_bushels(c, settlements)
        remaining = bushels - delivered
        status_label = _deliveries_fill_status_label(fill_status)

        group = groups.setdefault(
            (commodity, location),
            {"total_bushels": 0, "delivered": 0, "remaining": 0, "rows": []},
        )
        group["total_bushels"] += bushels
        group["delivered"] += delivered
        group["remaining"] += remaining
        group["rows"].append({
            "Contract #": c.contract_number or "",
            "Status": status_label,
            "_status_sort": _DELIVERIES_STATUS_SORT[status_label],
            "Price": f"${c.price:.2f}" if c.price is not None else "\u2014",
            "Bushels": f"{bushels:,}",
            "Delivered": f"{delivered:,}",
            "Remaining": f"{remaining:,}",
        })

    if not groups:
        st.info(
            f"No contracts deliver in {calendar.month_name[selected_month]} {selected_year}."
        )
        return

    ordered_keys = sorted(groups.keys())

    # ----- Summary -----
    st.markdown("#### Summary")
    summary_df = pd.DataFrame([
        {
            "Commodity": commodity,
            "Location": location,
            "Total Bushels": f"{groups[(commodity, location)]['total_bushels']:,}",
            "Delivered": f"{groups[(commodity, location)]['delivered']:,}",
            "Remaining": f"{groups[(commodity, location)]['remaining']:,}",
        }
        for (commodity, location) in ordered_keys
    ])
    st.dataframe(summary_df, hide_index=True, width='stretch')

    # A little whitespace before the details.
    st.write("")

    # ----- Details -----
    st.markdown("#### Details")
    # One cached bucket listing tells us which contracts have a PDF, so we only
    # render an "Open PDF" link for those rows (others stay blank).
    available = set(list_available_contract_numbers()) if storage_available() else set()
    for (commodity, location) in ordered_keys:
        group = groups[(commodity, location)]
        st.markdown(
            f"**{commodity} @ {location}** \u2014 "
            f"{group['total_bushels']:,} bu total, "
            f"{group['delivered']:,} delivered, "
            f"{group['remaining']:,} remaining"
        )
        detail_df = (
            pd.DataFrame(group["rows"])
            .sort_values(["_status_sort", "Contract #"])
            .drop(columns=["_status_sort"])
            .reset_index(drop=True)
        )
        detail_df["PDF"] = [
            signed_url_for_contract(cn) if cn in available else None
            for cn in detail_df["Contract #"]
        ]
        st.dataframe(
            detail_df,
            hide_index=True,
            width='stretch',
            column_config={
                "PDF": st.column_config.LinkColumn("PDF", display_text="Open PDF"),
            },
        )


def main():
    """Main dashboard application."""
    
    # Title
    st.title("🌾 Bushel Management Dashboard")
    st.markdown("---")
    
    # Show database path (for debugging)
    user_db_path = Path(DB_PATH)
    
    with st.sidebar.expander("ℹ️ Settings", expanded=True):  # Expanded by default for debugging
        st.write(f"**Environment:** {'Streamlit Cloud' if is_streamlit_cloud() else 'Local' if not is_colab() else 'Colab'}")
        st.write(f"**Database Path:** `{user_db_path}`")
        st.write(f"**Project Path:** `{PROJECT_PATH}`")
        st.write(f"**Script Location:** `{Path(__file__).absolute()}`")
        
        # Always show file listings for debugging
        data_folder = Path(DB_PATH).parent
        st.write(f"\n**Data folder path:** `{data_folder}`")
        st.write(f"**Data folder exists:** {'✅ Yes' if data_folder.exists() else '❌ No'}")
        
        if data_folder.exists():
            st.write(f"\n**Files in data folder:**")
            try:
                files = list(data_folder.iterdir())
                if files:
                    for f in sorted(files):
                        if f.is_file():
                            size_kb = f.stat().st_size / 1024
                            st.write(f"  📄 {f.name} ({size_kb:.1f} KB)")
                        else:
                            st.write(f"  📁 {f.name}/")
                else:
                    st.write("  (empty folder)")
            except Exception as e:
                st.write(f"  ❌ Error: {e}")
        else:
            st.write(f"  ❌ Data folder does not exist!")
        
        # Also check the project root to see what's there
        st.write(f"\n**Project root:** `{PROJECT_PATH}`")
        st.write(f"**Project root exists:** {'✅ Yes' if Path(PROJECT_PATH).exists() else '❌ No'}")
        try:
            root_files = list(Path(PROJECT_PATH).iterdir())
            st.write(f"**Files/folders in project root (first 20):**")
            for f in sorted(root_files)[:20]:
                if f.is_dir():
                    st.write(f"  📁 {f.name}/")
                else:
                    size_kb = f.stat().st_size / 1024
                    st.write(f"  📄 {f.name} ({size_kb:.1f} KB)")
        except Exception as e:
            st.write(f"  ❌ Error: {e}")
        
        # Check if database file exists
        file_exists_os = os.path.exists(DB_PATH)
        file_exists_path = Path(DB_PATH).exists()
        file_size = os.path.getsize(DB_PATH) / 1024 if file_exists_os else 0
        
        if file_exists_os and file_exists_path:
            st.success(f"✅ Database file found!")
            st.write(f"**File size:** {file_size:.1f} KB")
            st.write(f"**os.path.exists():** ✅")
            st.write(f"**Path().exists():** ✅")
        else:
            st.error(f"⚠️ Database file not found!")
            st.write(f"**Looking for:** `{DB_PATH}`")
            st.write(f"**os.path.exists():** {'✅' if file_exists_os else '❌'}")
            st.write(f"**Path().exists():** {'✅' if file_exists_path else '❌'}")
            if file_exists_os != file_exists_path:
                st.warning("⚠️ Path check mismatch! This might indicate a permission or path resolution issue.")
            st.info("""
            **To fix:**
            1. Check if file is in GitHub: https://github.com/dgableman/bushel-management-dashboard/tree/main/data
            2. If missing, add it and push to GitHub
            3. Reboot Streamlit Cloud app (⋮ menu → Reboot app)
            """)
        
        # Add a button to clear cache and retry
        if st.button("🔄 Clear Cache & Retry Connection"):
            st.cache_resource.clear()
            st.rerun()
    
    # Get database session
    db = get_database_session()
    if db is None:
        st.stop()
    
    # Get all contracts and settlements for calculations
    all_contracts = get_all_contracts(db)
    all_settlements = get_all_settlements(db)
    
    # Tabs for different views
    tab_deliveries, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["🚚 Deliveries", "🌾 Crop Year Sales", "📅 Deliveries by Month", "📦 Bins", "📦 Bins 3D", "📋 Contracts", "📥 Export"])
    
    with tab_deliveries:
        render_deliveries_tab(db, all_contracts, all_settlements)
    
    with tab1:
        st.subheader("Crop Year Sales")
        
        # Initialize drill-down session state
        if 'drilldown_crop' not in st.session_state:
            st.session_state.drilldown_crop = None
        if 'drilldown_status' not in st.session_state:
            st.session_state.drilldown_status = None
        if 'drilldown_status' not in st.session_state:
            st.session_state.drilldown_status = None
        
        # Crop year selector
        current_crop_year = get_current_crop_year()
        crop_year_options = get_display_year_options(current_crop_year)
        selected_crop_year = st.selectbox(
            "Crop Year",
            options=crop_year_options,
            index=crop_year_options.index(current_crop_year) if current_crop_year in crop_year_options else len(crop_year_options) - 1,
            format_func=lambda x: f"{x} (Oct 1, {x} - Sep 30, {x+1})"
        )
        
        # Clear drill-down if crop year changes
        if 'last_crop_year' not in st.session_state:
            st.session_state.last_crop_year = selected_crop_year
        elif st.session_state.last_crop_year != selected_crop_year:
            st.session_state.drilldown_crop = None
            st.session_state.drilldown_status = None
            st.session_state.last_crop_year = selected_crop_year
        
        start_date, end_date = get_crop_year_date_range(selected_crop_year)
        st.caption(f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
        
        # Calculate sales data
        sales_data = calculate_crop_year_sales(db, selected_crop_year)
        
        # Debug: Show raw data
        if st.checkbox("🔍 Show raw sales data", key="debug_sales_data"):
            st.write("**Sales Data:**")
            st.json(sales_data)
        
        if not sales_data:
            st.info("No data found for the selected crop year.")
        else:
            # Crop price inputs
            st.markdown("### Crop Prices - estimate for current price")
            crop_prices = {}
            crops = sorted(sales_data.keys())
            
            # Get default prices from market data (yfinance)
            default_prices = get_market_prices()
            
            price_cols = st.columns(min(4, len(crops)))
            for idx, crop in enumerate(crops):
                with price_cols[idx % 4]:
                    default_price = default_prices.get(crop, 0.0)
                    crop_prices[crop] = st.number_input(
                        f"{crop} Price ($/bu)",
                        min_value=0.0,
                        value=default_price,
                        step=0.01,
                        key=f"price_{crop}_{selected_crop_year}"
                    )
            
            # Calculate open revenue with prices
            for crop in crops:
                if crop in crop_prices:
                    sales_data[crop]['open_revenue'] = sales_data[crop]['open_bushels'] * crop_prices[crop]
            
            st.markdown("---")
            
            # Revenue Chart
            revenue_data = []
            total_sold_revenue = 0.0
            total_contracted_revenue = 0.0
            total_open_revenue = 0.0
            total_sold_bushels = 0
            total_contracted_bushels = 0
            total_open_bushels = 0
            
            for crop in crops:
                data = sales_data[crop]
                sold_rev = data['sold_revenue']
                contracted_rev = data['contracted_revenue']
                open_rev = data['open_revenue']
                sold_bu = data['sold_bushels']
                contracted_bu = data['contracted_bushels']
                open_bu = data['open_bushels']
                
                revenue_data.append({
                    'Crop': crop,
                    'Sold': sold_rev,
                    'Contracted': contracted_rev,
                    'Open': open_rev,
                    'Sold_Bushels': sold_bu,
                    'Contracted_Bushels': contracted_bu,
                    'Open_Bushels': open_bu
                })
                
                total_sold_revenue += sold_rev
                total_contracted_revenue += contracted_rev
                total_open_revenue += open_rev
                total_sold_bushels += sold_bu
                total_contracted_bushels += contracted_bu
                total_open_bushels += open_bu
            
            if revenue_data:
                # Add total row to the list (will be last)
                revenue_data.append({
                    'Crop': 'TOTAL',
                    'Sold': total_sold_revenue,
                    'Contracted': total_contracted_revenue,
                    'Open': total_open_revenue,
                    'Sold_Bushels': total_sold_bushels,
                    'Contracted_Bushels': total_contracted_bushels,
                    'Open_Bushels': total_open_bushels
                })
                
                # Create dataframe with crops first, then TOTAL
                df_revenue = pd.DataFrame(revenue_data)
                df_revenue = df_revenue.set_index('Crop')
                
                # Calculate totals for each row
                df_revenue['Total'] = df_revenue['Sold'] + df_revenue['Contracted'] + df_revenue['Open']
                
                # Calculate average prices per bushel
                df_revenue['Avg_Price_Sold'] = df_revenue.apply(
                    lambda row: row['Sold'] / row['Sold_Bushels'] if row['Sold_Bushels'] > 0 else 0.0, axis=1
                )
                df_revenue['Avg_Price_Contracted'] = df_revenue.apply(
                    lambda row: row['Contracted'] / row['Contracted_Bushels'] if row['Contracted_Bushels'] > 0 else 0.0, axis=1
                )
                df_revenue['Avg_Price_Open'] = df_revenue.apply(
                    lambda row: row['Open'] / row['Open_Bushels'] if row['Open_Bushels'] > 0 else 0.0, axis=1
                )
                
                # Calculate percentages for each segment (relative to total for that crop)
                df_revenue['Pct_Sold'] = df_revenue.apply(
                    lambda row: (row['Sold'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                df_revenue['Pct_Contracted'] = df_revenue.apply(
                    lambda row: (row['Contracted'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                df_revenue['Pct_Open'] = df_revenue.apply(
                    lambda row: (row['Open'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                
                # Format percentage text (only show if >= 1% to avoid cluttering)
                # Round to whole numbers and format as "xx%"
                df_revenue['Text_Sold'] = df_revenue.apply(
                    lambda row: f"{int(round(row['Pct_Sold']))}%" if row['Pct_Sold'] >= 1.0 else "", axis=1
                )
                df_revenue['Text_Contracted'] = df_revenue.apply(
                    lambda row: f"{int(round(row['Pct_Contracted']))}%" if row['Pct_Contracted'] >= 1.0 else "", axis=1
                )
                df_revenue['Text_Open'] = df_revenue.apply(
                    lambda row: f"{int(round(row['Pct_Open']))}%" if row['Pct_Open'] >= 1.0 else "", axis=1
                )
                
                # Create stacked horizontal bar chart - add in correct order: Settled, Contracted, Open
                fig_revenue = go.Figure()
                
                # Add Settled first (leftmost in bar, first in legend)
                fig_revenue.add_trace(go.Bar(
                    name='Settled',
                    y=df_revenue.index,
                    x=df_revenue['Sold'],
                    orientation='h',
                    marker_color='#2ecc71',
                    customdata=df_revenue[['Sold_Bushels', 'Avg_Price_Sold']].values,
                    text=df_revenue['Text_Sold'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='black', size=18, family='Arial Black'),
                    hovertemplate='Settled: $%{x:,.0f}<br>Bushels: %{customdata[0]:,.0f}<br>Avg Price: $%{customdata[1]:.2f}/bu<br>━━━━━━━━━━━━━━━━<extra></extra>',
                    legendrank=1,
                    showlegend=True
                ))
                # Add Contracted second (middle in bar, second in legend)
                fig_revenue.add_trace(go.Bar(
                    name='Contracted',
                    y=df_revenue.index,
                    x=df_revenue['Contracted'],
                    orientation='h',
                    marker_color='#3498db',
                    customdata=df_revenue[['Contracted_Bushels', 'Avg_Price_Contracted']].values,
                    text=df_revenue['Text_Contracted'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='white', size=18, family='Arial Black'),
                    hovertemplate='Contracted: $%{x:,.0f}<br>Bushels: %{customdata[0]:,.0f}<br>Avg Price: $%{customdata[1]:.2f}/bu<br>━━━━━━━━━━━━━━━━<extra></extra>',
                    legendrank=2,
                    showlegend=True
                ))
                # Calculate total average price (Total Revenue / Total Bushels) for hover
                df_revenue['Total_Bushels'] = df_revenue['Sold_Bushels'] + df_revenue['Contracted_Bushels'] + df_revenue['Open_Bushels']
                df_revenue['Avg_Price_Total'] = df_revenue.apply(
                    lambda row: row['Total'] / row['Total_Bushels'] if row['Total_Bushels'] > 0 else 0.0, axis=1
                )
                
                # Add Open last (rightmost in bar, last in legend, shows total at end)
                fig_revenue.add_trace(go.Bar(
                    name='Open',
                    y=df_revenue.index,
                    x=df_revenue['Open'],
                    orientation='h',
                    marker_color='#e74c3c',
                    customdata=df_revenue[['Total', 'Open_Bushels', 'Avg_Price_Open', 'Total_Bushels', 'Avg_Price_Total']].values,
                    text=df_revenue['Text_Open'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='white', size=18, family='Arial Black'),
                    hovertemplate='Open: $%{x:,.0f}<br>Bushels: %{customdata[1]:,.0f}<br>Avg Price: $%{customdata[2]:.2f}/bu<br>━━━━━━━━━━━━━━━━<br>Total: $%{customdata[0]:,.0f}<br>Total Bushels: %{customdata[3]:,.0f}<br>Avg Price: $%{customdata[4]:.2f}/bu<extra></extra>',
                    legendrank=3,
                    showlegend=True
                ))
                
                # Reverse the order so TOTAL appears at bottom
                category_array = list(df_revenue.index)
                category_array.reverse()  # Reverse so TOTAL (last) appears at bottom
                
                fig_revenue.update_layout(
                    barmode='stack',
                    title='Revenue',
                    xaxis_title='Revenue ($)',
                    yaxis_title='',  # Remove crop label
                    height=max(400, (len(crops) + 1) * 50),  # +1 for total row
                    hovermode='y unified',
                    hoverlabel=dict(
                        bgcolor='white',
                        bordercolor='black',
                        font_size=12,
                        namelength=-1
                    ),
                    yaxis=dict(categoryorder='array', categoryarray=category_array, showticklabels=True),
                    legend=dict(
                        traceorder='normal',
                        itemclick='toggle',
                        itemdoubleclick=False
                    ),
                    clickmode='event+select'
                )
                # Debug toggle (temporary)
                debug_selection = st.checkbox("🔍 Show selection debug info", key="debug_revenue_selection")
                
                # Handle chart selection - Streamlit stores selection in session state
                chart_revenue = st.plotly_chart(fig_revenue, width='stretch', on_select="rerun", key="revenue_chart")
                
                # Check for selection - prioritize session state key first
                # IMPORTANT: Check immediately after chart render, before any other processing
                selection_data = None
                selection_key_found = None
                
                # First, check the session state key directly (this is where Streamlit stores it)
                if 'revenue_chart' in st.session_state:
                    chart_state = st.session_state['revenue_chart']
                    if isinstance(chart_state, dict) and 'selection' in chart_state:
                        sel = chart_state['selection']
                        # Check if there are actual points selected (not empty)
                        if isinstance(sel, dict):
                            points_list = sel.get('points', [])
                            # Only process if points array is not empty
                            if points_list and len(points_list) > 0:
                                selection_data = sel
                                selection_key_found = 'revenue_chart'
                
                # Fallback: Check if chart_revenue return value has selection
                if selection_data is None and chart_revenue is not None:
                    if isinstance(chart_revenue, dict) and 'selection' in chart_revenue:
                        sel = chart_revenue['selection']
                        if isinstance(sel, dict):
                            points_list = sel.get('points', [])
                            if points_list and len(points_list) > 0:
                                selection_data = sel
                    elif hasattr(chart_revenue, 'selection'):
                        sel = chart_revenue.selection
                        if hasattr(sel, 'points') and sel.points and len(sel.points) > 0:
                            selection_data = sel
                
                if debug_selection:
                    st.write("**Debug Info:**")
                    st.write(f"- chart_revenue type: {type(chart_revenue)}")
                    st.write(f"- chart_revenue value: {chart_revenue}")
                    st.write(f"- session_state['revenue_chart']: {st.session_state.get('revenue_chart', 'NOT FOUND')}")
                    st.write(f"- selection_data: {selection_data}")
                    st.write(f"- selection_key_found: {selection_key_found}")
                    if selection_data and 'points' in selection_data:
                        st.write(f"- points count: {len(selection_data.get('points', []))}")
                        st.write(f"- points: {selection_data.get('points', [])}")
                
                # Process selection if found (only if points exist)
                if selection_data is not None:
                    # Extract points from selection data
                    points = None
                    if isinstance(selection_data, dict):
                        points = selection_data.get('points', [])
                    elif hasattr(selection_data, 'points'):
                        points = selection_data.points
                    elif isinstance(selection_data, list):
                        points = selection_data
                    
                    if points and len(points) > 0:
                        # Get crop name from the first point (all points at same y-position have same crop name)
                        crop_name = None
                        if isinstance(points[0], dict):
                            crop_name = points[0].get('y') or points[0].get('label')
                        else:
                            crop_name = getattr(points[0], 'y', None) or getattr(points[0], 'label', None)
                        
                        if debug_selection:
                            st.write(f"**Processing selection:**")
                            st.write(f"- Total points selected: {len(points)}")
                            st.write(f"- crop_name: {crop_name}")
                            st.write(f"- All points: {points}")
                        
                        # Ignore clicks on TOTAL - clear selection and do nothing else
                        if isinstance(crop_name, str) and crop_name == 'TOTAL':
                            if 'revenue_chart' in st.session_state:
                                del st.session_state['revenue_chart']
                            # Don't process further for TOTAL
                        # Navigate to detail page for this crop (only if not TOTAL)
                        elif isinstance(crop_name, str) and crop_name != 'TOTAL':
                            st.session_state.drilldown_crop = crop_name
                            st.session_state.drilldown_status = None  # None means show all statuses
                            st.session_state.selected_crop_year = selected_crop_year  # Store for detail page
                            
                            # Clear selection AFTER processing to prevent re-triggering
                            if 'revenue_chart' in st.session_state:
                                del st.session_state['revenue_chart']
                            
                            # Navigate to detail page
                            st.switch_page("pages/crop_details.py")
            
            st.markdown("---")
            
            # Bushels Chart
            bushels_data = []
            total_sold_bushels = 0
            total_contracted_bushels = 0
            total_open_bushels = 0
            total_sold_revenue_bushels = 0.0
            total_contracted_revenue_bushels = 0.0
            total_open_revenue_bushels = 0.0
            
            for crop in crops:
                data = sales_data[crop]
                sold_bu = data['sold_bushels']
                contracted_bu = data['contracted_bushels']
                open_bu = data['open_bushels']
                sold_rev = data['sold_revenue']
                contracted_rev = data['contracted_revenue']
                open_rev = data['open_revenue']
                
                bushels_data.append({
                    'Crop': crop,
                    'Sold': sold_bu,
                    'Contracted': contracted_bu,
                    'Open': open_bu,
                    'Sold_Revenue': sold_rev,
                    'Contracted_Revenue': contracted_rev,
                    'Open_Revenue': open_rev
                })
                
                total_sold_bushels += sold_bu
                total_contracted_bushels += contracted_bu
                total_open_bushels += open_bu
                total_sold_revenue_bushels += sold_rev
                total_contracted_revenue_bushels += contracted_rev
                total_open_revenue_bushels += open_rev
            
            if bushels_data:
                # Add total row to the list (will be last)
                bushels_data.append({
                    'Crop': 'TOTAL',
                    'Sold': total_sold_bushels,
                    'Contracted': total_contracted_bushels,
                    'Open': total_open_bushels,
                    'Sold_Revenue': total_sold_revenue_bushels,
                    'Contracted_Revenue': total_contracted_revenue_bushels,
                    'Open_Revenue': total_open_revenue_bushels
                })
                
                # Create dataframe with crops first, then TOTAL
                df_bushels = pd.DataFrame(bushels_data)
                df_bushels = df_bushels.set_index('Crop')
                
                # Calculate totals for each row
                df_bushels['Total'] = df_bushels['Sold'] + df_bushels['Contracted'] + df_bushels['Open']
                
                # Calculate average prices per bushel
                df_bushels['Avg_Price_Sold'] = df_bushels.apply(
                    lambda row: row['Sold_Revenue'] / row['Sold'] if row['Sold'] > 0 else 0.0, axis=1
                )
                df_bushels['Avg_Price_Contracted'] = df_bushels.apply(
                    lambda row: row['Contracted_Revenue'] / row['Contracted'] if row['Contracted'] > 0 else 0.0, axis=1
                )
                df_bushels['Avg_Price_Open'] = df_bushels.apply(
                    lambda row: row['Open_Revenue'] / row['Open'] if row['Open'] > 0 else 0.0, axis=1
                )
                
                # Calculate percentages for each segment (relative to total for that crop)
                df_bushels['Pct_Sold'] = df_bushels.apply(
                    lambda row: (row['Sold'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                df_bushels['Pct_Contracted'] = df_bushels.apply(
                    lambda row: (row['Contracted'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                df_bushels['Pct_Open'] = df_bushels.apply(
                    lambda row: (row['Open'] / row['Total'] * 100) if row['Total'] > 0 else 0.0, axis=1
                )
                
                # Format percentage text (only show if >= 1% to avoid cluttering)
                # Round to whole numbers and format as "xx%"
                df_bushels['Text_Sold'] = df_bushels.apply(
                    lambda row: f"{int(round(row['Pct_Sold']))}%" if row['Pct_Sold'] >= 1.0 else "", axis=1
                )
                df_bushels['Text_Contracted'] = df_bushels.apply(
                    lambda row: f"{int(round(row['Pct_Contracted']))}%" if row['Pct_Contracted'] >= 1.0 else "", axis=1
                )
                df_bushels['Text_Open'] = df_bushels.apply(
                    lambda row: f"{int(round(row['Pct_Open']))}%" if row['Pct_Open'] >= 1.0 else "", axis=1
                )
                
                # Create stacked horizontal bar chart - add in correct order: Sold, Contracted, Open
                fig_bushels = go.Figure()
                
                # Add Sold first (leftmost in bar, first in legend)
                fig_bushels.add_trace(go.Bar(
                    name='Sold',
                    y=df_bushels.index,
                    x=df_bushels['Sold'],
                    orientation='h',
                    marker_color='#2ecc71',
                    customdata=df_bushels[['Sold_Revenue', 'Avg_Price_Sold']].values,
                    text=df_bushels['Text_Sold'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='black', size=18, family='Arial Black'),
                    hovertemplate='Sold: %{x:,.0f} bu<br>Revenue: $%{customdata[0]:,.0f}<br>Avg Price: $%{customdata[1]:.2f}/bu<br>━━━━━━━━━━━━━━━━<extra></extra>',
                    legendrank=1,
                    showlegend=True
                ))
                # Add Contracted second (middle in bar, second in legend)
                fig_bushels.add_trace(go.Bar(
                    name='Contracted',
                    y=df_bushels.index,
                    x=df_bushels['Contracted'],
                    orientation='h',
                    marker_color='#3498db',
                    customdata=df_bushels[['Contracted_Revenue', 'Avg_Price_Contracted']].values,
                    text=df_bushels['Text_Contracted'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='white', size=18, family='Arial Black'),
                    hovertemplate='Contracted: %{x:,.0f} bu<br>Revenue: $%{customdata[0]:,.0f}<br>Avg Price: $%{customdata[1]:.2f}/bu<br>━━━━━━━━━━━━━━━━<extra></extra>',
                    legendrank=2,
                    showlegend=True
                ))
                # Add Open last (rightmost in bar, last in legend, shows total at end)
                fig_bushels.add_trace(go.Bar(
                    name='Open',
                    y=df_bushels.index,
                    x=df_bushels['Open'],
                    orientation='h',
                    marker_color='#e74c3c',
                    customdata=df_bushels[['Total', 'Open_Revenue', 'Avg_Price_Open']].values,
                    text=df_bushels['Text_Open'],
                    textposition='inside',
                    insidetextanchor='middle',
                    insidetextfont=dict(color='white', size=18, family='Arial Black'),
                    hovertemplate='Open: %{x:,.0f} bu<br>Revenue: $%{customdata[1]:,.0f}<br>Avg Price: $%{customdata[2]:.2f}/bu<br>━━━━━━━━━━━━━━━━<br><b>Total: %{customdata[0]:,.0f} bu</b><extra></extra>',
                    legendrank=3,
                    showlegend=True
                ))
                
                # Reverse the order so TOTAL appears at bottom
                category_array = list(df_bushels.index)
                category_array.reverse()  # Reverse so TOTAL (last) appears at bottom
                
                fig_bushels.update_layout(
                    barmode='stack',
                    title='Bushels',
                    xaxis_title='Bushels',
                    yaxis_title='',  # Remove crop label
                    height=max(400, (len(crops) + 1) * 50),  # +1 for total row
                    hovermode='y unified',
                    hoverlabel=dict(
                        bgcolor='white',
                        bordercolor='black',
                        font_size=12,
                        namelength=-1
                    ),
                    yaxis=dict(categoryorder='array', categoryarray=category_array, showticklabels=True),
                    legend=dict(
                        traceorder='normal',
                        itemclick='toggle',
                        itemdoubleclick=False
                    ),
                    clickmode='event+select'
                )
                # Debug toggle (temporary)
                debug_selection_bushels = st.checkbox("🔍 Show selection debug info", key="debug_bushels_selection")
                
                # Handle chart selection - Streamlit stores selection in session state
                chart_bushels = st.plotly_chart(fig_bushels, width='stretch', on_select="rerun", key="bushels_chart")
                
                # Check for selection - try multiple approaches
                selection_data = None
                selection_key_found = None
                
                # Approach 1: Check if chart_bushels return value is the selection
                if chart_bushels is not None:
                    if isinstance(chart_bushels, dict):
                        if 'selection' in chart_bushels:
                            selection_data = chart_bushels['selection']
                        elif 'points' in chart_bushels:
                            selection_data = chart_bushels
                    elif hasattr(chart_bushels, 'selection'):
                        selection_data = chart_bushels.selection
                    elif hasattr(chart_bushels, 'points'):
                        selection_data = chart_bushels
                
                # Approach 2: Check session state for selection keys
                if selection_data is None:
                    all_keys = list(st.session_state.keys())
                    # Try common key patterns
                    possible_keys = [
                        "bushels_chart.selection",
                        "bushels_chart_selection", 
                        "bushels_chart",
                        f"bushels_chart.selection.{selected_crop_year}"
                    ]
                    
                    for key in possible_keys:
                        if key in st.session_state:
                            val = st.session_state[key]
                            if val is not None and (hasattr(val, 'points') or isinstance(val, (dict, list))):
                                selection_data = val
                                selection_key_found = key
                                break
                    
                    # Also check all keys containing 'bushels' and 'selection'
                    if selection_data is None:
                        for key in all_keys:
                            if 'bushels_chart' in key.lower() and 'selection' in key.lower():
                                val = st.session_state[key]
                                if val is not None:
                                    selection_data = val
                                    selection_key_found = key
                                    break
                
                if debug_selection_bushels:
                    st.write("**Debug Info:**")
                    st.write(f"- chart_bushels type: {type(chart_bushels)}")
                    st.write(f"- chart_bushels value: {chart_bushels}")
                    st.write(f"- selection_data: {selection_data}")
                    st.write(f"- selection_key_found: {selection_key_found}")
                    st.write("**All session state keys:**")
                    for key in sorted(st.session_state.keys()):
                        if 'bushels' in key.lower() or 'selection' in key.lower():
                            st.write(f"- `{key}`: {st.session_state[key]}")
                
                # Also check if chart_bushels return value has selection
                if selection_data is None:
                    if hasattr(chart_bushels, 'selection'):
                        selection_data = chart_bushels.selection
                    elif isinstance(chart_bushels, dict) and 'selection' in chart_bushels:
                        selection_data = chart_bushels['selection']
                    elif chart_bushels is not None:
                        # chart_bushels might be the selection data itself
                        if hasattr(chart_bushels, 'points') or isinstance(chart_bushels, (list, dict)):
                            selection_data = chart_bushels
                
                # Process selection if found
                if selection_data is not None:
                    # Extract points from selection data
                    points = None
                    if isinstance(selection_data, dict):
                        points = selection_data.get('points', [])
                    elif hasattr(selection_data, 'points'):
                        points = selection_data.points
                    elif isinstance(selection_data, list):
                        points = selection_data
                    
                    if points and len(points) > 0:
                        # Get crop name from the first point (all points at same y-position have same crop name)
                        crop_name = None
                        if isinstance(points[0], dict):
                            crop_name = points[0].get('y') or points[0].get('label')
                        else:
                            crop_name = getattr(points[0], 'y', None) or getattr(points[0], 'label', None)
                        
                        # Ignore clicks on TOTAL - clear selection and do nothing else
                        if isinstance(crop_name, str) and crop_name == 'TOTAL':
                            if 'bushels_chart' in st.session_state:
                                del st.session_state['bushels_chart']
                            # Don't process further for TOTAL - exit early
                            pass
                        # Navigate to detail page for this crop (only if not TOTAL)
                        elif isinstance(crop_name, str) and crop_name != 'TOTAL':
                            st.session_state.drilldown_crop = crop_name
                            st.session_state.drilldown_status = None  # None means show all statuses
                            st.session_state.selected_crop_year = selected_crop_year  # Store for detail page
                            
                            # Clear selection AFTER processing to prevent re-triggering
                            if 'bushels_chart' in st.session_state:
                                del st.session_state['bushels_chart']
                            
                            # Navigate to detail page
                            st.switch_page("pages/crop_details.py")
            
    
    with tab2:
        st.subheader("Deliveries by Month")
        
        # Crop year selector (same as Crop Year Sales tab)
        current_crop_year = get_current_crop_year()
        deliveries_crop_year_options = get_display_year_options(current_crop_year)
        selected_crop_year = st.selectbox(
            "Crop Year",
            options=deliveries_crop_year_options,
            index=(
                deliveries_crop_year_options.index(current_crop_year)
                if current_crop_year in deliveries_crop_year_options
                else len(deliveries_crop_year_options) - 1
            ),
            format_func=lambda x: f"{x} (Oct 1, {x} - Sep 30, {x+1})",
            key="deliveries_crop_year"
        )
        
        start_date, end_date = get_crop_year_date_range(selected_crop_year)
        st.caption(f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}")
        
        # Calculate monthly deliveries data
        monthly_data = calculate_monthly_deliveries(
            db, 
            selected_crop_year,
            all_contracts=all_contracts,
            all_settlements=all_settlements
        )
        
        if not monthly_data:
            st.info("No data found for the selected crop year.")
        else:
            # Find the first month with data across all crops
            all_months = set()
            for crop_data in monthly_data.values():
                all_months.update(crop_data.keys())
            
            if not all_months:
                st.info("No monthly data available.")
            else:
                first_month = min(all_months)
                months_to_show = sorted([m for m in all_months if m >= first_month])
                
                # Get all month names in crop year order for proper x-axis ordering
                all_month_names_in_order = [get_month_name_for_crop_year(m) for m in months_to_show]
                
                # Create chart data
                crops = sorted(monthly_data.keys())
                
                # Create figure with dual y-axis
                fig = go.Figure()
                
                # Prepare data for each crop
                for crop in crops:
                    crop_months = monthly_data[crop]
                    
                    # Only include months from first_month onwards
                    months = [m for m in months_to_show if m in crop_months]
                    
                    if not months:
                        continue
                    
                    # Get month names for x-axis (in crop year order)
                    month_names = [get_month_name_for_crop_year(m) for m in months]
                    
                    # Get prices and bushels
                    prices = [crop_months[m]['price'] for m in months]
                    bushels = [crop_months[m]['bushels'] for m in months]
                    
                    # Add price line with diamond markers (left y-axis)
                    fig.add_trace(go.Scatter(
                        x=month_names,
                        y=prices,
                        mode='lines+markers',
                        marker=dict(symbol='diamond', size=10),
                        name=f'{crop} Price',
                        yaxis='y',
                        line=dict(width=2),
                        hovertemplate=f'<b>{crop} Price</b><br>Month: %{{x}}<br>Price: $%{{y:.2f}}/bu<extra></extra>'
                    ))
                    
                    # Add bushels line with square markers (right y-axis)
                    fig.add_trace(go.Scatter(
                        x=month_names,
                        y=bushels,
                        mode='lines+markers',
                        marker=dict(symbol='square', size=10),
                        name=f'{crop} Bushels',
                        yaxis='y2',
                        line=dict(width=2, dash='dot'),
                        hovertemplate=f'<b>{crop} Bushels</b><br>Month: %{{x}}<br>Bushels: %{{y:,.0f}} bu<extra></extra>'
                    ))
                
                # Update layout with dual y-axes - ensure months stay in crop year order
                fig.update_layout(
                    title='Deliveries by Month',
                    xaxis=dict(
                        title='Month',
                        tickangle=-45 if len(months_to_show) > 6 else 0,
                        categoryorder='array',
                        categoryarray=all_month_names_in_order
                    ),
                    yaxis=dict(
                        title='Price ($/bu)',
                        side='left',
                        showgrid=True
                    ),
                    yaxis2=dict(
                        title='Bushels',
                        side='right',
                        overlaying='y',
                        showgrid=False
                    ),
                    hovermode='x unified',
                    height=600,
                    legend=dict(
                        traceorder='normal',
                        yanchor="top",
                        y=0.99,
                        xanchor="left",
                        x=1.01
                    )
                )
                
                st.plotly_chart(fig, width='stretch')
                
                # Cumulative Deliveries Chart
                st.markdown("### Cumulative Deliveries by Month")
                
                # Create cumulative chart
                fig_cumulative = go.Figure()
                
                # Prepare cumulative data for each crop
                for crop in crops:
                    crop_months = monthly_data[crop]
                    
                    # Only include months from first_month onwards
                    months = [m for m in months_to_show if m in crop_months]
                    
                    if not months:
                        continue
                    
                    # Get month names for x-axis (in crop year order)
                    month_names = [get_month_name_for_crop_year(m) for m in months]
                    
                    # Calculate cumulative bushels
                    cumulative_bushels = []
                    running_total = 0
                    for m in months:
                        running_total += crop_months[m]['bushels']
                        cumulative_bushels.append(running_total)
                    
                    # Add cumulative line
                    fig_cumulative.add_trace(go.Scatter(
                        x=month_names,
                        y=cumulative_bushels,
                        mode='lines+markers',
                        marker=dict(size=10),
                        name=f'{crop} Cumulative',
                        line=dict(width=2),
                        hovertemplate=f'<b>{crop} Cumulative</b><br>Month: %{{x}}<br>Cumulative Bushels: %{{y:,.0f}} bu<extra></extra>'
                    ))
                
                # Update layout for cumulative chart
                fig_cumulative.update_layout(
                    title='Cumulative Deliveries by Month',
                    xaxis=dict(
                        title='Month',
                        tickangle=-45 if len(months_to_show) > 6 else 0,
                        categoryorder='array',
                        categoryarray=all_month_names_in_order
                    ),
                    yaxis=dict(
                        title='Cumulative Bushels',
                        showgrid=True
                    ),
                    hovermode='x unified',
                    height=500,
                    legend=dict(
                        traceorder='normal',
                        yanchor="top",
                        y=0.99,
                        xanchor="left",
                        x=1.01
                    )
                )
                
                st.plotly_chart(fig_cumulative, width='stretch')
                
                # Show summary data
                with st.expander("📊 View Monthly Data Summary"):
                    for crop in crops:
                        st.markdown(f"**{crop}**")
                        crop_months = monthly_data[crop]
                        
                        summary_data = []
                        for month_num in sorted(crop_months.keys()):
                            if month_num >= first_month:
                                data = crop_months[month_num]
                                summary_data.append({
                                    'Month': get_month_name_for_crop_year(month_num),
                                    'Bushels': f"{data['bushels']:,.0f}",
                                    'Gross Amount': f"${data['gross_amount']:,.2f}",
                                    'Price': f"${data['price']:.2f}/bu"
                                })
                        
                        if summary_data:
                            df_summary = pd.DataFrame(summary_data)
                            st.dataframe(df_summary, width='stretch', hide_index=True)
                            st.markdown("---")
    
    with tab3:
        st.subheader("Bins")
        
        # Crop year selector (same as other tabs)
        current_crop_year = get_current_crop_year()
        bins_crop_year_options = get_display_year_options(current_crop_year)
        selected_crop_year = st.selectbox(
            "Crop Year",
            options=bins_crop_year_options,
            index=(
                bins_crop_year_options.index(current_crop_year)
                if current_crop_year in bins_crop_year_options
                else len(bins_crop_year_options) - 1
            ),
            format_func=lambda x: f"{x} (Oct 1, {x} - Sep 30, {x+1})",
            key="bins_crop_year"
        )
        
        # View mode selector (radio button) and include empty bins checkbox
        col1, col2 = st.columns([2, 1])
        with col1:
            view_mode = st.radio(
                "View Mode",
                options=["View by Crop", "View by Location"],
                index=0,  # Default to "View by Crop"
                key="bins_view_mode",
                horizontal=True
            )
        with col2:
            include_empty_bins = st.checkbox(
                "Include empty bins",
                value=False,
                key="bins_include_empty",
                help="If checked, shows all bins regardless of storage. If unchecked, only shows bins with storage for the selected crop year."
            )
        
        # Get bins grouped by crop or location based on view mode
        try:
            if view_mode == "View by Crop":
                bins_by_group = get_bins_with_storage_by_crop(db, selected_crop_year, include_empty=include_empty_bins)
                group_label = "Crop"
            else:
                # Import here to avoid import errors if function doesn't exist (handles caching issues)
                from reports.bin_queries import get_bins_with_storage_by_location
                bins_by_group = get_bins_with_storage_by_location(db, selected_crop_year, include_empty=include_empty_bins)
                group_label = "Location"
        except Exception as e:
            st.error(f"Error loading bin data: {e}")
            import traceback
            with st.expander("Error details"):
                st.code(traceback.format_exc())
            bins_by_group = {}
            group_label = "Crop"

        open_contract_allocation = build_open_contract_allocation_by_bin(db, selected_crop_year)
        
        # Debug: Show what groups were found
        if st.checkbox("🔍 Debug: Show bin data", key="debug_bins"):
            st.write(f"**View Mode:** {view_mode}")
            st.write(f"**{group_label}s found:** {sorted(list(bins_by_group.keys()))}")
            st.write(f"**Total {group_label}s:** {len(bins_by_group)}")
            
            # Also show all bin_names and crop_storage for comparison
            if view_mode == "View by Location":
                from reports.bin_queries import get_all_bin_names, get_crop_storage_for_year
                all_bin_names = get_all_bin_names(db)
                all_storage = get_crop_storage_for_year(db, selected_crop_year)
                st.write(f"**All bin_names:** {len(all_bin_names)} bins")
                st.write(f"**All crop_storage records for year {selected_crop_year}:** {len(all_storage)} records")
                
                # Show unique locations from bin_names
                unique_locations_from_bins = set()
                for bn in all_bin_names:
                    if bn.location:
                        unique_locations_from_bins.add(bn.location)
                st.write(f"**Unique locations in bin_names:** {sorted(list(unique_locations_from_bins))}")
                
                # Show unique locations from crop_storage
                unique_locations_from_storage = set()
                for cs in all_storage:
                    if cs.location:
                        unique_locations_from_storage.add(cs.location)
                st.write(f"**Unique locations in crop_storage:** {sorted(list(unique_locations_from_storage))}")
            
            for group, bins_list in bins_by_group.items():
                st.write(f"**{group}:** {len(bins_list)} bins")
                for bin_name, crop_storage in bins_list:
                    current_content = crop_storage.current_content if crop_storage and hasattr(crop_storage, 'current_content') else 0
                    st.write(f"  - {bin_name.location} - {bin_name.bin_name}: capacity={bin_name.capacity}, current={current_content}")
        
        if not bins_by_group:
            if include_empty_bins:
                st.info("No bins found in the database.")
            else:
                st.info("No bins with storage found for the selected crop year. Try enabling 'Include empty bins' to see all bins.")
        else:
            # Create a chart for each group (crop or location)
            groups = sorted(bins_by_group.keys())
            
            for group in groups:
                st.markdown(f"### {group}")
                
                group_bins = bins_by_group[group]
                
                # Prepare data for stacked bar chart
                bin_labels = []
                settled_storage = []
                contracted_storage = []
                uncontracted_storage = []
                empty_storage = []
                empty_colors = []
                current_storage = []
                reference_heights = []
                availability_labels = []
                metrics_list = []
                
                for bin_name, crop_storage in sorted(group_bins, key=lambda b: (b[0].location, b[0].bin_name)):
                    if view_mode == "View by Location":
                        if crop_storage and hasattr(crop_storage, 'crop') and crop_storage.crop:
                            bin_label = f"{bin_name.bin_name} ({crop_storage.crop})"
                        else:
                            bin_label = f"{bin_name.bin_name} (Empty)"
                    else:
                        bin_label = f"{bin_name.location} - {bin_name.bin_name}"
                    bin_labels.append(bin_label)
                    
                    metrics = get_bin_storage_metrics(
                        crop_storage,
                        bin_name,
                        open_contract_bushels=open_contract_allocation.get(
                            (
                                (crop_storage.location or "") if crop_storage else "",
                                (crop_storage.bin_name or "") if crop_storage else "",
                                (crop_storage.crop or "") if crop_storage else "",
                            ),
                            0,
                        ),
                    )
                    metrics_list.append(metrics)
                    settled_storage.append(metrics['chart_settled'])
                    contracted_storage.append(metrics['chart_contracted'])
                    uncontracted_storage.append(metrics['chart_not_sold'])
                    empty_storage.append(metrics['chart_empty'])
                    empty_colors.append(
                        '#d5d8dc'
                    )
                    current_storage.append(metrics['current'])
                    reference_heights.append(metrics['reference'])
                    availability_labels.append(metrics['availability_label'])
                
                # Calculate bins per row - aim for ~2 inches per bin (assuming ~12 inch wide screen)
                # Approximately 4-5 bins per row to allow for ~2 inch width each
                bins_per_row = 4
                num_rows = (len(bin_labels) + bins_per_row - 1) // bins_per_row  # Ceiling division
                
                # Uniform bar width - Parameters that determine bin width:
                # 1. uniform_bar_width: Relative width of each bar (0.0-1.0, where 1.0 = full category spacing)
                #    - Smaller values = narrower bars
                #    - Set in the width parameter of go.Bar()
                # 2. bargap: Gap between bars (0.0-1.0, where 1.0 = full category spacing between bars)
                #    - Larger values = more space between bars = narrower appearance
                #    - Set in fig.update_layout(bargap=...)
                # 3. Number of bins: More bins = less space per bin (automatic)
                # Adjust based on number of bins to prevent fat bars
                num_bins = len(bin_labels)
                if num_bins == 1:
                    uniform_bar_width = 0.2  # Very narrow for single bin
                elif num_bins == 2:
                    uniform_bar_width = 0.35
                elif num_bins <= 4:
                    uniform_bar_width = 0.4
                else:
                    uniform_bar_width = 0.5  # Wider for more bins
                
                # Determine colors based on crop (need to get crop from first bin in group)
                # When viewing by location, bins can have different crops, so use crop from first bin
                first_storage = group_bins[0][1] if group_bins and group_bins[0][1] else None
                first_crop = first_storage.crop if first_storage and hasattr(first_storage, 'crop') else None
                crop_for_color = group.lower() if view_mode == "View by Crop" else (first_crop.lower() if first_crop else '')
                
                if crop_for_color == 'corn':
                    current_color = '#FFD700'  # Yellow
                    current_line_color = '#FFA500'  # Darker yellow/orange for border
                elif crop_for_color == 'soybeans':
                    current_color = '#8B4513'  # Brown
                    current_line_color = '#654321'  # Darker brown for border
                else:
                    current_color = '#3498db'  # Default blue
                    current_line_color = '#2980b9'
                
                # When viewing by location, bins may have different crops, so color each bin based on its crop
                # We'll need to handle this differently - use a list of colors for each bin
                if view_mode == "View by Location":
                    bin_colors = []
                    bin_line_colors = []
                    for bin_name, crop_storage in sorted(group_bins, key=lambda b: (b[0].location, b[0].bin_name)):
                        if crop_storage and hasattr(crop_storage, 'crop'):
                            crop_name = (crop_storage.crop or '').lower()
                        else:
                            crop_name = ''  # Empty bin
                        if crop_name == 'corn':
                            bin_colors.append('#FFD700')  # Yellow
                            bin_line_colors.append('#FFA500')
                        elif crop_name == 'soybeans':
                            bin_colors.append('#8B4513')  # Brown
                            bin_line_colors.append('#654321')
                        else:
                            bin_colors.append('#3498db')  # Default blue (or gray for empty)
                            bin_line_colors.append('#2980b9')
                else:
                    bin_colors = [current_color] * len(bin_labels)
                    bin_line_colors = [current_line_color] * len(bin_labels)
                
                def _add_bar_top_annotations(
                    fig_bins, *, use_subplot_grid=False, row_num=1, col_num=1, label_slice=None
                ):
                    indices = label_slice if label_slice is not None else range(len(bin_labels))
                    max_name_yshift = 95
                    for i in indices:
                        total_y = reference_heights[i]
                        metrics_html = _bin_metrics_label_text(metrics_list[i])
                        name_yshift, metrics_yshift = _bin_bar_label_yshifts(metrics_html)
                        max_name_yshift = max(max_name_yshift, name_yshift)
                        name_ann = dict(
                            x=bin_labels[i],
                            y=total_y,
                            text=_short_bin_title(bin_labels[i]),
                            showarrow=False,
                            yanchor="bottom",
                            yshift=name_yshift,
                            xanchor="center",
                            font=dict(size=15, color="black", family="Arial Black"),
                        )
                        metrics_ann = dict(
                            x=bin_labels[i],
                            y=total_y,
                            text=metrics_html,
                            showarrow=False,
                            yanchor="bottom",
                            yshift=metrics_yshift,
                            xanchor="center",
                            font=dict(size=13, color="black", family="Arial Black"),
                        )
                        for ann in (name_ann, metrics_ann):
                            if use_subplot_grid:
                                fig_bins.add_annotation(row=row_num, col=col_num, **ann)
                            else:
                                fig_bins.add_annotation(**ann)
                    return max_name_yshift + 45
                
                # Create subplots if multiple rows needed, otherwise single figure
                if num_rows > 1:
                    # Create subplots with num_rows rows
                    fig_bins = make_subplots(
                        rows=num_rows,
                        cols=1,
                        subplot_titles=None,  # No row titles - cleaner look
                        vertical_spacing=0.15,
                        shared_yaxes=True
                    )
                    
                    # Add bars for each row
                    chart_top_margin = 140
                    for row_idx in range(num_rows):
                        start_idx = row_idx * bins_per_row
                        end_idx = min(start_idx + bins_per_row, len(bin_labels))
                        
                        if start_idx < len(bin_labels):
                            row_bin_labels = bin_labels[start_idx:end_idx]
                            row_colors = bin_colors[start_idx:end_idx] if isinstance(bin_colors, list) else current_color
                            row_line_colors = bin_line_colors[start_idx:end_idx] if isinstance(bin_line_colors, list) else current_line_color
                            add_bins_stacked_bar_traces(
                                fig_bins,
                                row_bin_labels,
                                settled_storage[start_idx:end_idx],
                                contracted_storage[start_idx:end_idx],
                                uncontracted_storage[start_idx:end_idx],
                                empty_storage[start_idx:end_idx],
                                row_colors,
                                row_line_colors,
                                empty_colors[start_idx:end_idx],
                                uniform_bar_width,
                                show_legend=(row_idx == 0),
                                subplot_row=row_idx + 1,
                                subplot_col=1,
                            )
                            chart_top_margin = max(
                                chart_top_margin,
                                _add_bar_top_annotations(
                                    fig_bins,
                                    use_subplot_grid=True,
                                    row_num=row_idx + 1,
                                    col_num=1,
                                    label_slice=range(start_idx, end_idx),
                                ),
                            )
                    
                    # Update layout
                    fig_bins.update_layout(
                        title=f'{group} - Bin Storage Capacity',
                        barmode='stack',
                        hovermode='x unified',
                        height=350 * num_rows,  # Adjust height based on number of rows
                        margin=dict(t=chart_top_margin, b=20),
                        legend=dict(
                            traceorder='normal',
                            yanchor="top",
                            y=0.99,
                            xanchor="left",
                            x=1.01
                        ),
                        bargap=0.4 if len(bin_labels) <= 4 else 0.2  # More gap for fewer bins to prevent fat bars
                    )
                    
                    # Update x-axes for all subplots
                    for row_idx in range(num_rows):
                        start_idx = row_idx * bins_per_row
                        end_idx = min(start_idx + bins_per_row, len(bin_labels))
                        row_bin_labels = bin_labels[start_idx:end_idx] if start_idx < len(bin_labels) else []
                        
                        fig_bins.update_xaxes(
                            title='',
                            tickmode='array',
                            tickvals=row_bin_labels,
                            ticktext=[''] * len(row_bin_labels),
                            showticklabels=False,
                            row=row_idx+1, col=1
                        )
                    
                    # Y-axis: bushels cannot be negative (allow zoom, block pan below zero)
                    fig_bins.update_yaxes(
                        title='Bushels',
                        showgrid=True,
                        rangemode='nonnegative',
                        minallowed=0,
                        autorangeoptions_minallowed=0,
                    )
                    
                else:
                    # Single row - create regular figure
                    fig_bins = go.Figure()
                    
                    bar_colors = bin_colors if isinstance(bin_colors, list) else current_color
                    bar_line_colors = bin_line_colors if isinstance(bin_line_colors, list) else current_line_color
                    add_bins_stacked_bar_traces(
                        fig_bins,
                        bin_labels,
                        settled_storage,
                        contracted_storage,
                        uncontracted_storage,
                        empty_storage,
                        bar_colors,
                        bar_line_colors,
                        empty_colors,
                        uniform_bar_width,
                        show_legend=True,
                    )
                    
                    # Update layout - for single or few bins, use larger gap to prevent fat bars
                    num_bins = len(bin_labels)
                    gap_size = 0.6 if num_bins == 1 else (0.5 if num_bins == 2 else (0.4 if num_bins <= 4 else 0.2))
                    
                    fig_bins.update_layout(
                        title=f'{group} - Bin Storage Capacity',
                        xaxis=dict(
                            title='',
                            tickmode='array',
                            tickvals=bin_labels,
                            ticktext=[''] * len(bin_labels),
                            showticklabels=False,
                        ),
                        yaxis=dict(
                            title='Bushels',
                            showgrid=True,
                            rangemode='nonnegative',
                            minallowed=0,
                            autorangeoptions=dict(minallowed=0),
                        ),
                        barmode='stack',
                        hovermode='x unified',
                        height=500,
                        legend=dict(
                            traceorder='normal',
                            yanchor="top",
                            y=0.99,
                            xanchor="left",
                            x=1.01
                        ),
                        bargap=gap_size,  # Larger gap for fewer bins to prevent fat bars
                    )
                    chart_top_margin = _add_bar_top_annotations(fig_bins)
                    fig_bins.update_layout(margin=dict(t=chart_top_margin, b=20))
                
                st.plotly_chart(fig_bins, width='stretch')
                _render_bin_availability_captions(availability_labels)
                
                # Summary table for this group
                with st.expander(f"📊 View {group} Bin Details"):
                    summary_data = []
                    for bin_name, crop_storage in sorted(group_bins, key=lambda b: (b[0].location, b[0].bin_name)):
                        m = get_bin_storage_metrics(
                            crop_storage,
                            bin_name,
                            open_contract_bushels=open_contract_allocation.get(
                                (
                                    (crop_storage.location or "") if crop_storage else "",
                                    (crop_storage.bin_name or "") if crop_storage else "",
                                    (crop_storage.crop or "") if crop_storage else "",
                                ),
                                0,
                            ),
                        )
                        row_data = {
                            'Location': bin_name.location or 'N/A',
                            'Bin Name': bin_name.bin_name or 'N/A',
                            'Initial (bu)': f"{m['initial']:,.0f}",
                            'Current (bu)': f"{m['current']:,.0f}",
                            'Settled (bu)': f"{m['settled']:,.0f}",
                            'Open Contracts (bu)': f"{m.get('open_contracts', m['contracted']):,.0f}",
                            'Not Sold (bu)': f"{m.get('not_sold', m['available_to_market']):,.0f}",
                            'Total Capacity (bu)': f"{m['capacity']:,.0f}",
                            'Preferred Crop': bin_name.preferred_crop if hasattr(bin_name, 'preferred_crop') else 'N/A',
                            'Load Status': crop_storage.load_status if crop_storage and hasattr(crop_storage, 'load_status') else 'N/A'
                        }
                        # Add actual crop when viewing by location (since bins can have different crops)
                        if view_mode == "View by Location":
                            row_data['Crop'] = crop_storage.crop if crop_storage and hasattr(crop_storage, 'crop') else 'Empty'
                        summary_data.append(row_data)
                    
                    if summary_data:
                        df_bins = pd.DataFrame(summary_data)
                        st.dataframe(df_bins, width='stretch', hide_index=True)
                
                st.markdown("---")
    
    with tab4:
        st.subheader("Bins 3D")
        
        # Crop year selector
        current_crop_year = get_current_crop_year()
        bins2_crop_year_options = get_display_year_options(current_crop_year)
        selected_crop_year = st.selectbox(
            "Crop Year",
            options=bins2_crop_year_options,
            index=(
                bins2_crop_year_options.index(current_crop_year)
                if current_crop_year in bins2_crop_year_options
                else len(bins2_crop_year_options) - 1
            ),
            format_func=lambda x: f"{x} (Oct 1, {x} - Sep 30, {x+1})",
            key="bins2_crop_year"
        )
        
        # Get bins grouped by crop
        bins_by_crop = get_bins_with_storage_by_crop(db, selected_crop_year)
        open_contract_allocation_3d = build_open_contract_allocation_by_bin(db, selected_crop_year)
        
        if not bins_by_crop:
            st.info("No bins with storage found for the selected crop year.")
        else:
            # Crop color mapper
            crop_colors = {
                'Corn': 'gold',  # #FFD700
                'Soybeans': 'saddlebrown',  # #8B4513
            }
            
            # Function to generate cylinder using Surface (better approach)
            def cylinder(r, h, a=0, nt=100, nv=50):
                """
                Parametrize the cylinder of radius r, height h, base point a
                """
                theta = np.linspace(0, 2*np.pi, nt)
                v = np.linspace(a, a+h, nv)
                theta, v = np.meshgrid(theta, v)
                x = r * np.cos(theta)
                y = r * np.sin(theta)
                z = v
                return x, y, z
            
            def boundary_circle(r, h, nt=100):
                """
                r - boundary circle radius
                h - height above xOy-plane where the circle is included
                returns the circle parameterization
                """
                theta = np.linspace(0, 2*np.pi, nt)
                x = r * np.cos(theta)
                y = r * np.sin(theta)
                z = h * np.ones(theta.shape)
                return x, y, z
            
            # Create a chart for each crop
            crops = sorted(bins_by_crop.keys())
            
            for crop in crops:
                st.markdown(f"### {crop}")
                
                crop_bins = bins_by_crop[crop]
                
                # Prepare data
                bin_labels = []
                settled_storage_list = []
                contracted_storage_list = []
                uncontracted_storage_list = []
                empty_storage_list = []
                empty_color_list = []
                current_storage_list = []
                reference_heights_list = []
                availability_labels_list = []
                metrics_list_3d = []
                
                for bin_name, crop_storage in sorted(crop_bins, key=lambda b: (b[0].location, b[0].bin_name)):
                    bin_label = f"{bin_name.location} - {bin_name.bin_name}"
                    bin_labels.append(bin_label)
                    metrics = get_bin_storage_metrics(
                        crop_storage,
                        bin_name,
                        open_contract_bushels=open_contract_allocation_3d.get(
                            (
                                (crop_storage.location or "") if crop_storage else "",
                                (crop_storage.bin_name or "") if crop_storage else "",
                                (crop_storage.crop or "") if crop_storage else "",
                            ),
                            0,
                        ),
                    )
                    metrics_list_3d.append(metrics)
                    settled_storage_list.append(metrics['chart_settled'])
                    contracted_storage_list.append(metrics['chart_contracted'])
                    uncontracted_storage_list.append(metrics['chart_not_sold'])
                    empty_storage_list.append(metrics['chart_empty'])
                    empty_color_list.append(
                        '#d5d8dc'
                    )
                    current_storage_list.append(metrics['current'])
                    reference_heights_list.append(metrics['reference'])
                    availability_labels_list.append(metrics['availability_label'])
                
                # Build the 3D figure (one scene per bin when multiple — same view as single-bin)
                radius = 1.0
                num_bins = len(bin_labels)
                
                max_reference = max(reference_heights_list) if reference_heights_list else 1.0
                scale_factor = 20.0 / max_reference if max_reference > 0 else 0.001
                
                stored_color = crop_colors.get(crop, 'sienna')
                
                if num_bins <= 1:
                    fig = go.Figure()
                else:
                    subplot_titles = []
                    for lbl in bin_labels:
                        title = lbl.split(' - ', 1)[-1] if ' - ' in lbl else lbl
                        subplot_titles.append(title[:28] + ('…' if len(title) > 28 else ''))
                    fig = make_subplots(
                        rows=1,
                        cols=num_bins,
                        specs=[[{'type': 'scene'}] * num_bins],
                        subplot_titles=subplot_titles,
                        horizontal_spacing=0.04,
                    )
                
                settled_added = False
                contracted_added = False
                uncontracted_added = False
                empty_added = False
                label_pad_top = 1.8

                def _add_trace(trace):
                    if num_bins <= 1:
                        fig.add_trace(trace)
                    else:
                        fig.add_trace(trace, row=1, col=subplot_col)

                def _top_label_text(metrics):
                    return _bin_metrics_label_text(metrics)

                def _add_bin_top_label(stack_height, metrics):
                    z_top = stack_height + label_pad_top
                    _add_trace(go.Scatter3d(
                        x=[0], y=[0], z=[z_top],
                        mode="text",
                        text=[_top_label_text(metrics)],
                        textfont=dict(size=18, color="#000000", family="Arial Black"),
                        showlegend=False,
                        hoverinfo="skip",
                    ))

                def _add_cylinder_segment(height, z_base, color, legend_name, show_in_legend):
                    if height <= 0:
                        return z_base
                    x1, y1, z1 = cylinder(radius, height, a=z_base, nt=50, nv=30)
                    _add_trace(go.Surface(
                        x=x1, y=y1, z=z1,
                        colorscale=[[0, color], [1, color]],
                        showscale=False,
                        opacity=0.8,
                        name=legend_name,
                        showlegend=show_in_legend,
                    ))
                    xb_low, yb_low, zb_low = boundary_circle(radius, h=z_base, nt=50)
                    xb_up, yb_up, zb_up = boundary_circle(radius, h=z_base + height, nt=50)
                    _add_trace(go.Scatter3d(
                        x=xb_low.tolist() + [None] + xb_up.tolist(),
                        y=yb_low.tolist() + [None] + yb_up.tolist(),
                        z=zb_low.tolist() + [None] + zb_up.tolist(),
                        mode='lines',
                        line=dict(color=color, width=2),
                        opacity=0.9,
                        showlegend=False,
                    ))
                    return z_base + height

                for idx, bin_label in enumerate(bin_labels):
                    subplot_col = idx + 1
                    metrics = metrics_list_3d[idx]
                    settled_val = settled_storage_list[idx]
                    contracted_val = contracted_storage_list[idx]
                    uncontracted_val = uncontracted_storage_list[idx]
                    empty_val = empty_storage_list[idx]
                    empty_color = empty_color_list[idx]
                    
                    z_base = 0.0
                    if settled_val > 0:
                        z_base = _add_cylinder_segment(
                            settled_val * scale_factor, z_base, BIN_SETTLED_COLOR, 'Settled', not settled_added
                        )
                        settled_added = True
                    if contracted_val > 0:
                        z_base = _add_cylinder_segment(
                            contracted_val * scale_factor, z_base, BIN_CONTRACTED_COLOR, 'Open Contracts', not contracted_added
                        )
                        contracted_added = True
                    if uncontracted_val > 0:
                        z_base = _add_cylinder_segment(
                            uncontracted_val * scale_factor, z_base, stored_color, 'Not sold', not uncontracted_added
                        )
                        uncontracted_added = True
                    if empty_val > 0:
                        z_base = _add_cylinder_segment(
                            empty_val * scale_factor, z_base, empty_color, 'Empty', not empty_added
                        )
                        empty_added = True
                    
                    stack_height = max(reference_heights_list[idx] * scale_factor, 0.5)
                    _add_bin_top_label(stack_height, metrics)
                
                legend_cfg = dict(yanchor="top", y=0.99, xanchor="left", x=1.01)
                layout_margin = dict(l=10, r=10, t=45, b=0)
                if num_bins <= 1:
                    z_top = stack_height + label_pad_top + 0.5
                    fig.update_layout(
                        title=f'{crop} - 3D Bin Storage',
                        height=580,
                        margin=layout_margin,
                        legend=legend_cfg,
                        scene=dict(
                            xaxis=dict(visible=False),
                            yaxis=dict(visible=False),
                            zaxis=dict(range=[0, z_top], visible=False),
                            aspectmode='cube',
                            camera=dict(eye=dict(x=1.5, y=1.5, z=0.55)),
                        ),
                    )
                    fig.layout.scene.camera.projection.type = "orthographic"
                else:
                    max_stack_z = max(
                        reference_heights_list[i] * scale_factor
                        for i in range(num_bins)
                    ) if num_bins else 20.0
                    max_stack_z = max(max_stack_z, radius * 2) * 1.05 + label_pad_top
                    fig.update_scenes(
                        xaxis_visible=False,
                        yaxis_visible=False,
                        zaxis=dict(range=[0, max_stack_z], visible=False),
                        aspectmode='cube',
                        camera=dict(eye=dict(x=1.5, y=1.5, z=0.55)),
                    )
                    for scene_idx in range(num_bins):
                        scene_ref = fig.layout.scene if scene_idx == 0 else getattr(fig.layout, f'scene{scene_idx + 1}')
                        scene_ref.camera.projection.type = "orthographic"
                    fig.update_layout(
                        title=f'{crop} - 3D Bin Storage',
                        height=480,
                        margin=layout_margin,
                        legend=legend_cfg,
                    )
                
                st.plotly_chart(
                    fig,
                    width='stretch',
                    config={'scrollZoom': False},
                )
                _render_bin_availability_captions(availability_labels_list)
                st.markdown("---")
    
    with tab5:
        st.subheader("Contracts")
        
        # Get all contracts to determine available filter options
        all_contracts = get_all_contracts(db)
        
        if not all_contracts:
            st.info("No contracts found in the database.")
        else:
            year_basis = st.radio(
                "Contract year basis",
                options=["Crop Year", "Calendar Year"],
                horizontal=True,
                key="contract_year_basis",
                help=(
                    "Crop Year uses Oct 1 – Sep 30 (e.g. 2026 = Oct 2026 through Sep 2027). "
                    "Calendar Year uses Jan 1 – Dec 31. Choose one basis — not both."
                ),
            )
            crop_years = discover_contract_crop_years(all_contracts)
            calendar_years = discover_contract_calendar_years(all_contracts)
            
            # Crop types (normalized names)
            crop_types = set()
            for contract in all_contracts:
                if contract.commodity:
                    normalized = normalize_commodity_name(db, contract.commodity)
                    crop_types.add(normalized)
            crop_types = sorted(list(crop_types))
            
            # 4. Vendors (normalized buyer names)
            vendors = get_all_normalized_vendors(db, all_contracts)
            
            # 5. Fill statuses
            fill_statuses = ['None', 'Partial', 'Filled', 'Over']
            
            # Filter checkboxes
            st.markdown("### Filters")
            filter_col1, filter_col2, filter_col3, filter_col4, filter_col5 = st.columns(5)
            
            with filter_col1:
                if year_basis == "Crop Year":
                    st.markdown("**Crop Year**")
                    selected_year_filter = []
                    for cy in crop_years:
                        if st.checkbox(
                            f"{cy}",
                            key=f"contract_crop_year_{cy}",
                            value=True,
                        ):
                            selected_year_filter.append(cy)
                else:
                    st.markdown("**Calendar Year**")
                    selected_year_filter = []
                    for cal_y in calendar_years:
                        if st.checkbox(
                            f"{cal_y}",
                            key=f"contract_calendar_year_{cal_y}",
                            value=True,
                        ):
                            selected_year_filter.append(cal_y)
            
            delivery_months = delivery_months_for_year_selection(
                all_contracts, year_basis, selected_year_filter
            )
            with filter_col2:
                st.markdown("**Delivery Month**")
                if not selected_year_filter:
                    st.caption("Select at least one year to see delivery months.")
                    selected_delivery_months = []
                elif not delivery_months:
                    st.caption("No delivery months for the selected year(s).")
                    selected_delivery_months = []
                else:
                    selected_delivery_months = []
                    for dm in delivery_months:
                        if st.checkbox(
                            f"{dm}",
                            key=f"contract_delivery_month_{year_basis}_{dm}",
                            value=True,
                        ):
                            selected_delivery_months.append(dm)
            
            with filter_col3:
                st.markdown("**Crop Type**")
                selected_crop_types = []
                for ct in crop_types:
                    if st.checkbox(f"{ct}", key=f"contract_crop_type_{ct}", value=True):
                        selected_crop_types.append(ct)
            
            with filter_col4:
                st.markdown("**Vendor**")
                selected_vendors = []
                for vendor in vendors:
                    if st.checkbox(f"{vendor}", key=f"contract_vendor_{vendor}", value=True):
                        selected_vendors.append(vendor)
            
            with filter_col5:
                st.markdown("**Fill Status**")
                selected_fill_statuses = []
                for fs in fill_statuses:
                    if st.checkbox(f"{fs}", key=f"contract_fill_status_{fs}", value=True):
                        selected_fill_statuses.append(fs)
            
            # Filter contracts based on selections
            filtered_contracts = []
            for contract in all_contracts:
                if not contract_matches_year_basis(
                    contract.delivery_start, year_basis, selected_year_filter
                ):
                    continue
                
                if contract.delivery_start:
                    contract_month_key = format_delivery_month_key(contract.delivery_start)
                    if selected_delivery_months and contract_month_key not in selected_delivery_months:
                        continue
                elif selected_delivery_months:
                    continue
                
                # Filter by crop type
                if contract.commodity:
                    normalized_crop = normalize_commodity_name(db, contract.commodity)
                    if selected_crop_types and normalized_crop not in selected_crop_types:
                        continue
                elif selected_crop_types:
                    continue
                
                # Filter by vendor
                if contract.buyer_name:
                    normalized_vendor = normalize_vendor_name(db, contract.buyer_name)
                    if selected_vendors and normalized_vendor not in selected_vendors:
                        continue
                elif selected_vendors:
                    continue
                
                # Filter by fill status
                contract_fill_status = contract.fill_status or 'None'
                if selected_fill_statuses and contract_fill_status not in selected_fill_statuses:
                    continue
                
                filtered_contracts.append(contract)
            
            st.markdown("---")
            
            if not filtered_contracts:
                st.info("No contracts match the selected filters.")
            else:
                # Prepare data for stacked bar chart
                # Group by vendor + crop type combination and fill status
                chart_data = {}  # {vendor_crop_key: {fill_status: bushels}}
                
                for contract in filtered_contracts:
                    vendor = normalize_vendor_name(db, contract.buyer_name or 'Unknown')
                    crop_type = normalize_commodity_name(db, contract.commodity) if contract.commodity else 'Unknown'
                    fill_status = contract.fill_status or 'None'
                    bushels = contract.bushels or 0
                    
                    # Create combined key: "Vendor X (Corn)"
                    vendor_crop_key = f"{vendor} ({crop_type})"
                    
                    if vendor_crop_key not in chart_data:
                        chart_data[vendor_crop_key] = {}
                    if fill_status not in chart_data[vendor_crop_key]:
                        chart_data[vendor_crop_key][fill_status] = 0
                    chart_data[vendor_crop_key][fill_status] += bushels
                
                # Create stacked bar chart
                if chart_data:
                    vendor_crop_list = sorted(chart_data.keys())
                    fill_status_colors = {
                        'None': '#3498db',
                        'Partial': '#f39c12',
                        'Filled': '#2ecc71',
                        'Over': '#e74c3c'
                    }
                    
                    # Control bar width like bins tab
                    num_bars = len(vendor_crop_list)
                    if num_bars == 1:
                        uniform_bar_width = 0.2  # Very narrow for single bar
                    elif num_bars == 2:
                        uniform_bar_width = 0.35
                    elif num_bars <= 4:
                        uniform_bar_width = 0.4
                    else:
                        uniform_bar_width = 0.5  # Wider for more bars
                    
                    # Determine gap size
                    gap_size = 0.6 if num_bars == 1 else (0.5 if num_bars == 2 else (0.4 if num_bars <= 4 else 0.2))
                    
                    fig = go.Figure()
                    
                    # Add a trace for each fill status
                    for fill_status in fill_statuses:
                        if fill_status not in selected_fill_statuses:
                            continue
                        
                        bushels_by_vendor_crop = [chart_data.get(vendor_crop, {}).get(fill_status, 0) for vendor_crop in vendor_crop_list]
                        
                        fig.add_trace(go.Bar(
                            name=fill_status,
                            x=vendor_crop_list,
                            y=bushels_by_vendor_crop,
                            marker=dict(color=fill_status_colors.get(fill_status, '#95a5a6')),
                            hovertemplate='%{fullData.name}: %{y:,.0f} bu<extra></extra>',
                            width=[uniform_bar_width] * len(vendor_crop_list)  # Uniform width for all bars
                        ))
                    
                    fig.update_layout(
                        title='Bushels by Vendor and Crop Type (Stacked by Fill Status)',
                        xaxis_title='',
                        yaxis_title='Bushels',
                        barmode='stack',
                        hovermode='x unified',
                        height=500,
                        bargap=gap_size,  # Gap between bars
                        xaxis=dict(
                            tickangle=-45 if len(vendor_crop_list) > 5 else 0,
                            tickfont=dict(size=12, family='Arial Black', color='black')  # Bold and larger
                        )
                    )
                    
                    st.plotly_chart(fig, width='stretch')
                    st.markdown("---")
                
                # Display contracts table - split by crop type
                st.markdown("### Contract Details")
                
                # Group contracts by crop type
                contracts_by_crop = {}
                for contract in filtered_contracts:
                    contract_crop_year = None
                    if contract.delivery_start:
                        try:
                            contract_crop_year = get_crop_year_from_date(contract.delivery_start)
                        except Exception:
                            pass
                    
                    normalized_crop = normalize_commodity_name(db, contract.commodity) if contract.commodity else 'Unknown'
                    normalized_vendor = normalize_vendor_name(db, contract.buyer_name) if contract.buyer_name else 'Unknown'
                    
                    if normalized_crop not in contracts_by_crop:
                        contracts_by_crop[normalized_crop] = []
                    
                    contracts_by_crop[normalized_crop].append({
                        'Contract Number': contract.contract_number,
                        'Crop Year': contract_crop_year if contract_crop_year else '',
                        'Vendor': normalized_vendor,
                        'Bushels': contract.bushels or 0,
                        'Price ($/bu)': f"${contract.price:.2f}" if contract.price else '',
                        'Fill Status': contract.fill_status or 'None',
                        'Status': contract.status or 'Active',
                        'Date Sold': contract.date_sold.strftime('%Y-%m-%d') if contract.date_sold else '',
                        'Delivery Start': contract.delivery_start.strftime('%Y-%m-%d') if contract.delivery_start else '',
                        'Delivery End': contract.delivery_end.strftime('%Y-%m-%d') if contract.delivery_end else ''
                    })
                
                # Display separate table for each crop type
                for crop in sorted(contracts_by_crop.keys()):
                    st.markdown(f"#### {crop}")
                    df_contracts = pd.DataFrame(contracts_by_crop[crop])
                    st.dataframe(df_contracts, width='stretch', hide_index=True)
                    st.markdown("---")

                render_contract_pdf_picker(
                    [c.contract_number for c in filtered_contracts],
                    key_prefix="contracts_tab",
                )
    
    with tab6:
        st.subheader("Export Data")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📥 Export to Excel", width='stretch'):
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    
                    wb = Workbook()
                    
                    # Contracts sheet
                    if all_contracts:
                        ws_contracts = wb.active
                        ws_contracts.title = "Contracts"
                        headers = ['Contract #', 'Commodity', 'Bushels', 'Price', 'Basis', 'Status', 'Date Sold', 'Buyer']
                        ws_contracts.append(headers)
                        
                        # Style headers
                        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        for cell in ws_contracts[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        for c in all_contracts:
                            ws_contracts.append([
                                c.contract_number, normalize_commodity_name(db, c.commodity), c.bushels, c.price, 
                                c.basis, c.status, c.date_sold, c.buyer_name
                            ])
                    
                    # Settlements sheet
                    filtered_settlements = [s for s in all_settlements if s.status != 'Header']
                    if filtered_settlements:
                        ws_settlements = wb.create_sheet("Settlements")
                        headers = ['Settlement ID', 'Contract #', 'Bushels', 'Price', 'Date Delivered', 'Bin', 'Buyer', 'Gross Amount', 'Net Amount', 'Adjustments', 'Status']
                        ws_settlements.append(headers)
                        
                        # Style headers
                        for cell in ws_settlements[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        for s in filtered_settlements:
                            ws_settlements.append([
                                s.settlement_ID, s.contract_id, s.bushels, s.price,
                                s.date_delivered, s.bin, s.buyer, s.gross_amount,
                                s.net_amount, s.adjustments, s.status
                            ])
                    
                    output_path = f"{PROJECT_PATH}/bushel_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    wb.save(output_path)
                    st.success(f"✓ Excel exported: {output_path}")
                    
                    # Provide download link
                    with open(output_path, 'rb') as f:
                        st.download_button(
                            label="Download Excel File",
                            data=f.read(),
                            file_name=f"bushel_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Error exporting to Excel: {e}")
                    import traceback
                    with st.expander("Error details"):
                        st.code(traceback.format_exc())
        
        with col2:
            if st.button("📄 Export to CSV", width='stretch'):
                try:
                    # Create a combined CSV or separate files
                    st.write("**Export Options:**")
                    
                    # Contracts CSV
                    if all_contracts:
                        df_contracts = pd.DataFrame([{
                            'Contract #': c.contract_number,
                            'Commodity': normalize_commodity_name(db, c.commodity),
                            'Bushels': c.bushels,
                            'Price': c.price,
                            'Basis': c.basis,
                            'Status': c.status,
                            'Date Sold': c.date_sold,
                            'Buyer': c.buyer_name
                        } for c in all_contracts])
                        
                        csv_contracts = df_contracts.to_csv(index=False)
                        st.download_button(
                            label="Download Contracts CSV",
                            data=csv_contracts,
                            file_name=f"contracts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )
                    
                    # Settlements CSV
                    filtered_settlements = [s for s in all_settlements if s.status != 'Header']
                    if filtered_settlements:
                        df_settlements = pd.DataFrame([{
                            'Settlement ID': s.settlement_ID,
                            'Contract #': s.contract_id,
                            'Bushels': s.bushels,
                            'Price': s.price,
                            'Date Delivered': s.date_delivered,
                            'Bin': s.bin,
                            'Buyer': s.buyer,
                            'Gross Amount': s.gross_amount,
                            'Net Amount': s.net_amount,
                            'Adjustments': s.adjustments,
                            'Status': s.status
                        } for s in filtered_settlements])
                        
                        csv_settlements = df_settlements.to_csv(index=False)
                        st.download_button(
                            label="Download Settlements CSV",
                            data=csv_settlements,
                            file_name=f"settlements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )
                    
                    if not all_contracts and not filtered_settlements:
                        st.warning("No data to export.")
                except Exception as e:
                    st.error(f"Error exporting to CSV: {e}")
                    import traceback
                    with st.expander("Error details"):
                        st.code(traceback.format_exc())

if __name__ == "__main__":
    main()
